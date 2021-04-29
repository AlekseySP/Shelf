from kivy.lang import Builder
from kivy.core.window import Window
from kivy.uix.boxlayout import BoxLayout
from kivy.clock import Clock
from kivy.core.audio import SoundLoader
from kivy.properties import ObjectProperty
from kivymd.uix.filemanager import MDFileManager
from kivymd.app import MDApp
from kivymd.toast import toast
from openpyxl import load_workbook
import os

KV = '''
BoxLayout:
    orientation: 'vertical'
    padding: self.width*0.05 #20
    spacing: 5
    text_input: text_input
    place_label: place_label
    
    # canvas.before:
    #     Color:
    #         rgba: 0, 0, 0, 1
    #     Rectangle:
    #         pos: self.pos
    #         size: self.size
    
    BoxLayout:
        size_hint: .5, None
        height: '48dp'
        spacing: 5

        MDTextButton:
            text: 'Обновить базу адресов'
            on_release: app.file_manager_open()

    MDTextField:
        id: text_input
        hint_text: 'Введите артикул'
        helper_text: ''
        helper_text_mode: 'persistent'
        pos_hint: {'center_y': .8}
        on_text_validate: root.show_place()
        
    MDProgressBar:
        value: 50

    MDLabel:
        id: place_label
        text: 'A-00-00-00'
        halign: 'center'
        theme_text_color: 'Primary'
        # text_color: app.theme_cls.text_color
        font_size: self.width*0.17
        pos_hint: {'center_y': .4}
        # text_size: root.width, None

    BoxLayout:
        spacing: 5
        orientation: 'horizontal'

        # MDRectangleFlatButton:
        #     text: "Обновить"
        #     on_release: root.makeDict()
        #     elevation_normal: 20

        MDRectangleFlatButton:
            text: "Выход"
            #pos_hint: {'center_y': .5}
            on_release: app.stop()
'''


class MainScreen(BoxLayout):
    file_name = "shelfadress.txt"
    path = ""#"(/home/sveta/programming/git") # ("/")#('/data/media/0')
    file_path = os.path.join(path, file_name)
    if os.path.isfile(file_path) == False:
        innerAdressfile = open(file_name, "w+")
        innerAdressfile.write("TAB-057:A-01-05-03")
        innerAdressfile.close()

    # TODO добавить загрузку из xlsx файла для adres_dict
    adres_dict = {}
    loadfile = ObjectProperty(None)
    text_input = ObjectProperty(None)
    place_label = ObjectProperty(None)

    def show_keyboard(self, event):
        self.text_input.focus = True

    def show_place(self):
        self.beep2 = SoundLoader.load('b1.wav')
        self.beep1 = SoundLoader.load('b2.wav')
        self.actAdr = self.place_label.text
        self.adress = self.adres_dict
        self.inp = self.text_input.text.upper()
        if self.inp:
            if ";" in self.inp:
                self.out = list(self.inp.split(";"))[-1]
            else:
                self.out = list(self.inp.split())[-1]   
            if self.out in self.adress:
                self.place = self.adress[self.out]
                self.text_input.helper_text = self.out
                self.place_label.text=self.place
                if self.actAdr == self.place:
                	self.beep1.play()
                else:
                	self.beep2.play()
                Clock.schedule_once(self.show_keyboard)
            else:
                self.place_label.text= "Неизвестно"
                self.text_input.helper_text = ""
                self.beep2.play()
                Clock.schedule_once(self.show_keyboard)
        else:
            self.place_label.text="Нет ввода"
            self.text_input.helper_text = ""
            self.beep2.play()
        self.text_input.text = ""
        Clock.schedule_once(self.show_keyboard)



class ShelfApp(MDApp):
    def __init__(self, **kwargs):
        self.title = "Shelf 3"
        self.theme_cls.theme_style = "Dark"
        self.theme_cls.primary_palette = "Teal"
        super().__init__(**kwargs)
        Window.bind(on_keyboard=self.events)
        self.manager_open = False
        self.file_manager = MDFileManager(
            exit_manager=self.exit_manager,
            select_path=self.select_path,
            previous=False,
            )
        self.file_manager.ext = ['.xlsx']

    def build(self):
        #return MainScreen()
        return Builder.load_string(KV)

    def file_manager_open(self):
        self.file_manager.show("/")#('/data/media/0')
        self.manager_open = True
        # output manager to the screen

    def server_update(self, path):
        self.path = path
        self.delete_old_data()
        self.upload_data(self.path)

    def dict_from_xl(self, path):
        self.path = path
        self.wb = load_workbook(filename = self.path)
        self.ws = self.wb.worksheets[0]
        self.adres_dict = {}
        for i in range (1, len(self.ws["A"])):
            self.k = str(self.ws["A" + str(i)].value)
            self.v = str(self.ws["B" + str(i)].value)
            self.adres_dict.update([(self.k, self.v)])
        print("dictionary passed")
        return self.adres_dict

    def upload_data(self, path):
        # TODO функция заполнения внутреннего xlsx файла
        # из внешнего, указанного пользователем
        print("Загрузка завершена")
        

    def delete_old_data(self):
        # TODO функция очистки внутреннего файла перед
        # заполнением (возможно не нужна)
        print("В процессе...")


    def select_path(self, path):
        '''It will be called when you click on the file name
        or the catalog selection button.
        :type path: str;
        :param path: path to the selected directory or file;
        '''
        self.exit_manager()
        self.server_update(path)
        toast('Адреса обновлены')
    def exit_manager(self, *args):
        '''Called when the user reaches the root of the directory tree.'''
        self.manager_open = False
        self.file_manager.close()
    def events(self, instance, keyboard, keycode, text, modifiers):
        '''Called when buttons are pressed on the mobile device.'''
        if keyboard in (1001, 27):
            if self.manager_open:
                self.file_manager.back()
        return True


if __name__ == "__main__":
    ShelfApp().run()