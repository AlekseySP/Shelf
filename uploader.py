from openpyxl import load_workbook
import pymysql

Host = "remotemysql.com"
Port = 3306
User = "3ALraEDF4c"
Password = "L3pBenmNFz"

file = "Adr140920.xlsx"
wb = load_workbook(filename = file)
ws = wb.worksheets[0]
adres_dict = {}
for i in range (1, len(ws["A"])):
    k = str(ws["A" + str(i)].value)
    v = str(ws["B" + str(i)].value)
    adres_dict.update([(k, v)])
print("dictionary ready")


d = adres_dict
arts  = list(d.keys())
con = pymysql.connect(host=Host,
port=Port, user=User, passwd=Password,
db=User)
cur = con.cursor()
# sql = "UPDATE Articles set adress = %s where article = %s"
sql = "INSERT INTO Articles (Article, Adress) VALUES (%s, %s)"
# for i in arts:
#   art = i
 #  adr = d[i]
 #  val = (art, adr)
val = [(i, d[i]) for i in arts]
cur.executemany(sql, val)
con.commit()
cur.close()
con.close()
print("Data uploading complete.")